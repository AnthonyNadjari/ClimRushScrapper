"""
Merge all segment CSVs into one Excel + JSON summary.

Usage:
    python -m scraper.merge --input-dir artifacts/ --output-dir results/
"""

import argparse
import csv
import datetime
import glob
import json
import os
import sys

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from .models import Lead
from .utils import deduplicate, save_json_summary


DARK_BLUE = "1B2A4A"
ACCENT_BLUE = "2E5090"
LIGHT_BLUE = "D6E4F0"
LIGHTER_BLUE = "EBF1F8"
WHITE = "FFFFFF"

SEG_COLORS = {
    "creche": "27AE60",
    "hotel": "E67E22",
    "salle": "8E44AD",
    "restaurant": "E74C3C",
    "cabinet": "2980B9",
    "salon": "F39C12",
    "agence": "1ABC9C",
    "boulangerie": "D35400",
}

header_font = Font(name="Arial", bold=True, color=WHITE, size=11)
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
data_font = Font(name="Arial", size=10)
center_align = Alignment(horizontal="center", vertical="center")
left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
thin_border = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)
row_fill_even = PatternFill("solid", fgColor=LIGHTER_BLUE)
row_fill_odd = PatternFill("solid", fgColor=WHITE)

DISPLAY_COLS = {
    "nom_entreprise": "Entreprise",
    "adresse": "Adresse",
    "code_postal": "CP",
    "ville": "Ville",
    "telephone": "Telephone",
    "site_web": "Site Web",
    "note_google": "Note",
    "nb_avis": "Avis",
    "segment": "Segment",
    "source": "Source",
}

COL_WIDTHS = [40, 42, 8, 12, 16, 45, 7, 7, 25, 14]


def _get_seg_color(name: str) -> str:
    name_lower = name.lower()
    for key, color in SEG_COLORS.items():
        if key in name_lower:
            return color
    return ACCENT_BLUE


def load_csvs(input_dir: str) -> list[Lead]:
    """Load all segment CSVs from input directory (handles nested artifact dirs)."""
    all_leads = []
    patterns = [
        os.path.join(input_dir, "**", "segment_*.csv"),
        os.path.join(input_dir, "segment_*.csv"),
    ]
    files_found = set()
    for pattern in patterns:
        for path in glob.glob(pattern, recursive=True):
            if path in files_found:
                continue
            files_found.add(path)
            try:
                with open(path, "r", encoding="utf-8-sig") as f:
                    reader = csv.DictReader(f, delimiter=";")
                    for row in reader:
                        lead = Lead(**{k: row.get(k, "") for k in Lead.fieldnames()})
                        all_leads.append(lead)
                print(f"  Loaded: {path} ({sum(1 for _ in open(path, encoding='utf-8-sig'))-1} rows)")
            except Exception as e:
                print(f"  Error loading {path}: {e}")
    return all_leads


def _write_data_sheet(ws, leads: list[Lead], color: str = DARK_BLUE):
    """Write a data sheet with headers and formatted rows."""
    # Headers
    for c, (key, label) in enumerate(DISPLAY_COLS.items(), 1):
        cell = ws.cell(row=1, column=c, value=label)
        cell.font = header_font
        cell.fill = PatternFill("solid", fgColor=color)
        cell.alignment = header_align
        cell.border = thin_border

    # Data
    for r, lead in enumerate(leads, 2):
        fill = row_fill_even if r % 2 == 0 else row_fill_odd
        d = lead.to_dict()
        for c, key in enumerate(DISPLAY_COLS.keys(), 1):
            val = d.get(key, "")
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = data_font
            cell.fill = fill
            cell.border = thin_border
            if key in ("code_postal", "note_google", "nb_avis", "telephone"):
                cell.alignment = center_align
                if key == "telephone":
                    cell.number_format = "@"
            else:
                cell.alignment = left_align

    # Column widths
    for i, w in enumerate(COL_WIDTHS):
        from openpyxl.utils import get_column_letter
        ws.column_dimensions[get_column_letter(i + 1)].width = w

    ws.freeze_panes = "A2"
    if leads:
        ws.auto_filter.ref = f"A1:J{len(leads)+1}"


def build_excel(all_leads: list[Lead], output_path: str):
    """Build a professional Excel with dashboard + segment sheets."""
    valid = deduplicate([l for l in all_leads if l.is_valid()])

    # Group by segment
    segments = {}
    for lead in valid:
        seg = lead.segment or "Autres"
        if seg not in segments:
            segments[seg] = []
        segments[seg].append(lead)

    wb = Workbook()

    # ── Dashboard ──
    ws = wb.active
    ws.title = "Dashboard"
    ws.sheet_properties.tabColor = DARK_BLUE

    ws.merge_cells("A1:H2")
    c = ws["A1"]
    c.value = "CLIMRUSH - Leads B2B"
    c.font = Font(name="Arial", bold=True, color=WHITE, size=16)
    c.fill = PatternFill("solid", fgColor=DARK_BLUE)
    c.alignment = Alignment(horizontal="center", vertical="center")

    now = datetime.datetime.now().strftime("%d/%m/%Y %H:%M")
    ws.merge_cells("A3:H3")
    c = ws["A3"]
    c.value = f"Scraping multi-source : Google Maps + Pages Jaunes | {now}"
    c.font = Font(name="Arial", italic=True, color="999999", size=10)
    c.alignment = Alignment(horizontal="center", vertical="center")

    total = len(valid)
    tel_count = len([l for l in valid if l.telephone])
    site_count = len([l for l in valid if l.site_web])
    tel_pct = int(tel_count / total * 100) if total else 0
    site_pct = int(site_count / total * 100) if total else 0

    stat_big = Font(name="Arial", bold=True, color=DARK_BLUE, size=28)
    stat_label = Font(name="Arial", color="666666", size=10)

    for col, val, lbl in [("B", str(total), "Leads uniques"), ("D", str(tel_count), f"Telephones ({tel_pct}%)"), ("F", str(site_count), f"Sites web ({site_pct}%)")]:
        ws[f"{col}5"].value = val
        ws[f"{col}5"].font = stat_big
        ws[f"{col}5"].alignment = Alignment(horizontal="center", vertical="center")
        ws[f"{col}6"].value = lbl
        ws[f"{col}6"].font = stat_label
        ws[f"{col}6"].alignment = Alignment(horizontal="center", vertical="center")

    # Segment table
    seg_headers = ["Segment", "Leads", "Telephones", "Sites web", "% Tel", "% Site", "Sources"]
    for ci, h in enumerate(seg_headers, 1):
        cell = ws.cell(row=8, column=ci, value=h)
        cell.font = header_font
        cell.fill = PatternFill("solid", fgColor=ACCENT_BLUE)
        cell.alignment = header_align
        cell.border = thin_border

    row = 9
    for seg_name, seg_leads in segments.items():
        n = len(seg_leads)
        t = len([l for l in seg_leads if l.telephone])
        s = len([l for l in seg_leads if l.site_web])
        src = {}
        for l in seg_leads:
            src[l.source] = src.get(l.source, 0) + 1
        src_str = " + ".join(f"{v} {k}" for k, v in src.items())
        fill = row_fill_even if row % 2 == 0 else row_fill_odd
        vals = [seg_name, n, t, s, f"{int(t/n*100)}%" if n else "0%", f"{int(s/n*100)}%" if n else "0%", src_str]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=ci, value=v)
            cell.font = Font(name="Arial", bold=True, size=10) if ci == 1 else data_font
            cell.fill = fill
            cell.border = thin_border
            cell.alignment = center_align if ci > 1 else left_align
        row += 1

    # Total row
    fill_total = PatternFill("solid", fgColor=LIGHT_BLUE)
    for ci, v in enumerate(["TOTAL", total, tel_count, site_count, f"{tel_pct}%", f"{site_pct}%", ""], 1):
        cell = ws.cell(row=row, column=ci, value=v)
        cell.font = Font(name="Arial", bold=True, size=11, color=DARK_BLUE)
        cell.fill = fill_total
        cell.border = thin_border
        cell.alignment = center_align if ci > 1 else left_align

    for col, w in [("A", 25), ("B", 12), ("C", 14), ("D", 12), ("E", 10), ("F", 10), ("G", 30), ("H", 5)]:
        ws.column_dimensions[col].width = w

    # ── Tous les leads sheet ──
    ws_all = wb.create_sheet("Tous les leads")
    ws_all.sheet_properties.tabColor = ACCENT_BLUE
    _write_data_sheet(ws_all, valid, DARK_BLUE)

    # ── Per-segment sheets ──
    for seg_name, seg_leads in segments.items():
        safe = seg_name[:31]  # Excel sheet name max 31 chars
        ws_seg = wb.create_sheet(safe)
        ws_seg.sheet_properties.tabColor = _get_seg_color(seg_name)
        _write_data_sheet(ws_seg, seg_leads, _get_seg_color(seg_name))

    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    wb.save(output_path)
    print(f"  Excel saved: {output_path} ({total} leads)")
    return valid


def main():
    parser = argparse.ArgumentParser(description="Merge segment CSVs into Excel")
    parser.add_argument("--input-dir", required=True, help="Directory with segment CSVs")
    parser.add_argument("--output-dir", default="results/", help="Output directory")
    args = parser.parse_args()

    os.makedirs(args.output_dir, exist_ok=True)

    print("=" * 60)
    print("  CLIMRUSH MERGE — Consolidating segments")
    print("=" * 60)

    all_leads = load_csvs(args.input_dir)
    print(f"  Total loaded: {len(all_leads)} leads bruts")

    valid = deduplicate([l for l in all_leads if l.is_valid()])
    print(f"  After dedup: {len(valid)} leads uniques")

    # Build Excel
    xlsx_path = os.path.join(args.output_dir, "CLIMRUSH_Leads.xlsx")
    build_excel(all_leads, xlsx_path)

    # Save JSON summary
    json_path = os.path.join(args.output_dir, "latest.json")
    summary = save_json_summary(valid, json_path)
    print(f"  JSON saved: {json_path}")

    # Also save a combined CSV
    from .utils import save_csv
    csv_path = os.path.join(args.output_dir, "CLIMRUSH_Leads.csv")
    save_csv(all_leads, csv_path)
    print(f"  CSV saved: {csv_path}")

    print("=" * 60)
    print(f"  DONE: {len(valid)} leads | {summary['total_phones']} tel | {summary['total_websites']} sites")
    print("=" * 60)


if __name__ == "__main__":
    main()
